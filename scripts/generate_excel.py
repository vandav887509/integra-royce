#!/usr/bin/env python3
"""
generate_excel.py
=================
Generates one Excel file per machine (B21, B24, B25, B27) from
machine-data.json, matching the format of the reference files:
  - Charts sheet  (header info + 4 line charts)
  - Type 1 sheet  (DATE | ITEM | Data)
  - Type 2 sheet
  - Type 3 Short sheet
  - Type 3 Long sheet
  - Data sheet    (DATE | MACHINE | ITEM | Type | Data)

Usage:
    python3 scripts/generate_excel.py \\
        --json /var/www/integra-royce/dashboard/data/machine-data.json \\
        --out  /var/www/integra-royce/dashboard/data/excel

Output files:
    BOND_PULL_DATA_IGN2932M75_B21_bonder.xlsx
    BOND_PULL_DATA_IGN2932M75_B24_bonder.xlsx
    BOND_PULL_DATA_IGN2932M75_B25_bonder.xlsx
    BOND_PULL_DATA_IGN2932M75_B27_bonder.xlsx

Add to cron after process_csv.py:
    0 * * * * python3 /var/www/integra-royce/scripts/process_csv.py ... && \\
              python3 /var/www/integra-royce/scripts/generate_excel.py ...
"""

import json
import argparse
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl --break-system-packages")

PRODUCT = "IGN2932M75"

WIRE_SHEETS = [
    ("t1",  "Type 1",       "TYPE 1"),
    ("t2",  "Type 2",       "TYPE 2"),
    ("t3s", "Type 3 Short", "TYPE 3 SHORT"),
    ("t3l", "Type 3 Long",  "TYPE 3 LONG"),
]

HEADER_FONT = Font(name="Calibri", bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
NORMAL_FONT = Font(name="Calibri")
DATE_FMT    = "MM/DD/YY"


def make_header(ws, cols):
    for col, label in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def write_type_sheet(ws, machine_id, dates, values):
    """Write DATE | ITEM | Data rows for one wire type. Returns last data row."""
    make_header(ws, ["DATE", "ITEM", "Data"])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10

    row = 2
    for date_str, val in zip(dates, values):
        if val is None:
            continue
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            continue
        ws.cell(row=row, column=1, value=dt).number_format = DATE_FMT
        ws.cell(row=row, column=2, value=PRODUCT).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=val).font = NORMAL_FONT
        row += 1

    return row - 1  # last data row


def write_data_sheet(ws, machine_id, dates, t1, t2, t3s, t3l):
    """Write combined Data sheet: DATE | MACHINE | ITEM | Type | Data."""
    make_header(ws, ["DATE", "MACHINE", "ITEM", "Type", "Data"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10

    type_map = [
        ("TYPE 1",       t1),
        ("TYPE 2",       t2),
        ("TYPE 3 SHORT", t3s),
        ("TYPE 3 LONG",  t3l),
    ]

    row = 2
    for i, date_str in enumerate(dates):
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            continue
        for type_label, series in type_map:
            val = series[i] if i < len(series) else None
            if val is None:
                continue
            ws.cell(row=row, column=1, value=dt).number_format = DATE_FMT
            ws.cell(row=row, column=2, value=machine_id).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=PRODUCT).font = NORMAL_FONT
            ws.cell(row=row, column=4, value=type_label).font = NORMAL_FONT
            ws.cell(row=row, column=5, value=val).font = NORMAL_FONT
            row += 1


def add_line_chart(ws_charts, title, data_ws_name, data_rows, anchor):
    """Add a line chart to the Charts sheet."""
    chart         = LineChart()
    chart.title   = title
    chart.style   = 10
    chart.height  = 10
    chart.width   = 18
    chart.legend  = None

    if data_rows >= 2:
        data_ref = Reference(
            ws_charts.parent[data_ws_name],
            min_col=3, min_row=2, max_row=data_rows
        )
        chart.add_data(data_ref)

        cats = Reference(
            ws_charts.parent[data_ws_name],
            min_col=1, min_row=2, max_row=data_rows
        )
        chart.set_categories(cats)

        s = chart.series[0]
        s.graphicalProperties.line.solidFill = "4472C4"
        s.graphicalProperties.line.width     = 18000
        s.smooth = False

    ws_charts.add_chart(chart, anchor)


def generate_workbook(machine_id: str, data: dict, out_dir: Path):
    num_id = machine_id.replace("B", "")
    mdata  = data.get(num_id)
    if not mdata:
        print(f"  WARNING: no data for {machine_id}, skipping")
        return

    dates = mdata.get("dates", [])
    t1    = mdata.get("t1",  [])
    t2    = mdata.get("t2",  [])
    t3s   = mdata.get("t3s", [])
    t3l   = mdata.get("t3l", [])

    wb = Workbook()

    # Charts sheet
    ws_charts       = wb.active
    ws_charts.title = "Charts"
    ws_charts["C1"] = "MACHINE";   ws_charts["C1"].font = HEADER_FONT
    ws_charts["D1"] = machine_id;  ws_charts["D1"].font = HEADER_FONT
    ws_charts["E1"] = "ITEM NAME"; ws_charts["E1"].font = HEADER_FONT
    ws_charts["F1"] = PRODUCT;     ws_charts["F1"].font = HEADER_FONT
    ws_charts.column_dimensions["C"].width = 12
    ws_charts.column_dimensions["D"].width = 10
    ws_charts.column_dimensions["E"].width = 12
    ws_charts.column_dimensions["F"].width = 14

    # Type data sheets
    series_map = {"t1": t1, "t2": t2, "t3s": t3s, "t3l": t3l}
    last_rows  = {}

    for key, sheet_name, _ in WIRE_SHEETS:
        ws       = wb.create_sheet(sheet_name)
        last_row = write_type_sheet(ws, machine_id, dates, series_map[key])
        last_rows[sheet_name] = last_row

    # Data sheet
    write_data_sheet(wb.create_sheet("Data"), machine_id, dates, t1, t2, t3s, t3l)

    # Add 4 line charts to Charts sheet
    chart_configs = [
        (f"BOND DATA CHART :: TYPE 1",       "Type 1",       "B2"),
        (f"BOND DATA CHART :: TYPE 2",        "Type 2",       "K2"),
        (f"BOND DATA CHART :: TYPE 3 SHORT",  "Type 3 Short", "B22"),
        (f"BOND DATA CHART :: TYPE 3 LONG",   "Type 3 Long",  "K22"),
    ]
    for title, sheet_name, anchor in chart_configs:
        add_line_chart(ws_charts, title, sheet_name, last_rows.get(sheet_name, 1), anchor)

    fname = f"BOND_PULL_DATA_{PRODUCT}_{machine_id}_bonder.xlsx"
    fpath = out_dir / fname
    wb.save(fpath)
    print(f"  Saved: {fpath}  ({len(dates)} dates)")


def main():
    parser = argparse.ArgumentParser(description="Generate Excel files from machine-data.json")
    parser.add_argument("--json", default="/var/www/integra-royce/dashboard/data/machine-data.json")
    parser.add_argument("--out",  default="/var/www/integra-royce/dashboard/data/excel")
    args = parser.parse_args()

    json_path = Path(args.json)
    out_dir   = Path(args.out)

    if not json_path.exists():
        sys.exit(f"JSON not found: {json_path}")

    out_dir.mkdir(parents=True, exist_ok=True)

    with open(json_path) as f:
        data = json.load(f)

    print(f"Generating Excel files → {out_dir}")
    for machine_id in ["B21", "B24", "B25", "B27"]:
        generate_workbook(machine_id, data, out_dir)
    print("Done.")


if __name__ == "__main__":
    main()
